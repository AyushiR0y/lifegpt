from fastapi import APIRouter, File, UploadFile, Form, HTTPException, WebSocket
from fastapi.responses import FileResponse
import base64
import json
import re
import unicodedata
import torch
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM
from pdf2image import convert_from_bytes
from PIL import Image
import io
import os
from typing import List, Dict, Optional
import tempfile
import fitz  # PyMuPDF
import pdfplumber
import pytesseract
import glob
import urllib.error
import urllib.request
from fastapi.websockets import WebSocketDisconnect
from openpyxl import load_workbook
import xlrd

# Add after app initialization
class ConnectionManager:
    def __init__(self):
        self.active_connections: dict = {}
    
    async def connect(self, websocket: WebSocket, client_id: str):
        await websocket.accept()
        self.active_connections[client_id] = websocket
    
    def disconnect(self, client_id: str):
        if client_id in self.active_connections:
            del self.active_connections[client_id]
    
    async def send_message(self, message: str, client_id: str):
        if client_id in self.active_connections:
            try:
                await self.active_connections[client_id].send_text(message)
            except:
                pass

manager = ConnectionManager()
# Import transliteration rules
from transliteration import (
    TRANSLITERATE_WORDS,
    KEEP_ORIGINAL_WORDS,
    TRANSLITERATE_WORDS_LOWER,
    KEEP_ORIGINAL_WORDS_LOWER,
    LANGUAGE_CODE_TO_DICT_KEY,
    apply_translation_rules,
    get_transliteration,
    should_keep_original
)

router = APIRouter()

# Global model variables
text_model = None
text_tokenizer = None
device = None
FONTS = {}
MODEL_LOAD_ERROR = None
POPLLER_PATH = None


def load_env_file(env_path: Optional[str] = None) -> None:
    """Load key/value pairs from .env into process environment without overriding existing vars."""
    if env_path is None:
        env_path = os.path.join(os.path.dirname(__file__), ".env")

    if not os.path.exists(env_path):
        return

    try:
        with open(env_path, "r", encoding="utf-8") as env_file:
            for raw_line in env_file:
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue

                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception as e:
        print(f"Warning: unable to load .env file: {e}")


load_env_file()


PII_PATTERNS = [
    (re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"), "[EMAIL]"),
    (re.compile(r"\b\d{3}-\d{2}-\d{4}\b"), "[SSN]"),
    (re.compile(r"\b\d{4}\s?\d{4}\s?\d{4}\b"), "[AADHAAR]"),
    (re.compile(r"\b[A-Z]{5}\d{4}[A-Z]\b"), "[PAN]"),
    (re.compile(r"\b(?:\d[ -]*?){13,19}\b"), "[CARD_NUMBER]"),
    (re.compile(r"(?<!\d)(?:\+?\d{1,3}[-.\s]?)?(?:\(?\d{2,4}\)?[-.\s]?){2,4}\d{3,4}(?!\d)"), "[PHONE]"),
]


def mask_pii_text(text: str) -> str:
    if not text:
        return text

    masked = text
    for pattern, replacement in PII_PATTERNS:
        masked = pattern.sub(replacement, masked)
    return masked


def sanitize_chat_messages(messages: List[Dict]) -> List[Dict]:
    sanitized_messages = []
    for message in messages or []:
        sanitized_message = dict(message)
        content = sanitized_message.get("content")
        if isinstance(content, str):
            sanitized_message["content"] = mask_pii_text(content)
        elif isinstance(content, list):
            sanitized_parts = []
            for part in content:
                if isinstance(part, dict) and isinstance(part.get("text"), str):
                    sanitized_part = dict(part)
                    sanitized_part["text"] = mask_pii_text(sanitized_part["text"])
                    sanitized_parts.append(sanitized_part)
                else:
                    sanitized_parts.append(part)
            sanitized_message["content"] = sanitized_parts
        sanitized_messages.append(sanitized_message)

    return sanitized_messages


def tokenize_for_retrieval(text: str) -> List[str]:
    return [token for token in re.findall(r"[a-z0-9]+", (text or "").lower()) if len(token) > 2]


def chunk_text_for_retrieval(text: str, max_chars: int = 1600, overlap: int = 120) -> List[str]:
    clean_text = (text or "").strip()
    if not clean_text:
        return []

    if len(clean_text) <= max_chars:
        return [clean_text]

    chunks = []
    step = max(1, max_chars - overlap)
    for start in range(0, len(clean_text), step):
        chunk = clean_text[start:start + max_chars].strip()
        if chunk:
            chunks.append(chunk)
        if start + max_chars >= len(clean_text):
            break
    return chunks


def score_chunk(query_tokens: List[str], chunk_text: str) -> int:
    if not query_tokens:
        return 0

    chunk_tokens = set(tokenize_for_retrieval(chunk_text))
    if not chunk_tokens:
        return 0

    return sum(1 for token in set(query_tokens) if token in chunk_tokens)


def decode_data_url(data_url: str) -> tuple[bytes, str]:
    header, encoded = data_url.split(",", 1)
    mime_type = header[5:].split(";", 1)[0].lower()
    return base64.b64decode(encoded), mime_type


def extract_spreadsheet_sections(file_bytes: bytes, file_name: str) -> List[Dict]:
    sections: List[Dict] = []
    lower_name = file_name.lower()

    if lower_name.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            row_lines: List[str] = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(cell).strip() for cell in row if cell not in (None, "")]
                if not values:
                    continue
                row_lines.append(f"Row {row_index}: " + " | ".join(values))
            if row_lines:
                sections.append({"label": f"Sheet {sheet.title}", "text": "\n".join(row_lines)})
        return sections

    if lower_name.endswith(".xls"):
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        for sheet in workbook.sheets():
            row_lines = []
            for row_index in range(sheet.nrows):
                values = [str(sheet.cell_value(row_index, col_index)).strip() for col_index in range(sheet.ncols)]
                values = [value for value in values if value and value.lower() != "nan"]
                if not values:
                    continue
                row_lines.append(f"Row {row_index + 1}: " + " | ".join(values))
            if row_lines:
                sections.append({"label": f"Sheet {sheet.name}", "text": "\n".join(row_lines)})

    return sections


def extract_pdf_sections(pdf_bytes: bytes, max_pages: int = 6) -> List[Dict]:
    sections: List[Dict] = []

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_index, page in enumerate(pdf.pages[:max_pages], start=1):
                page_text = (page.extract_text() or "").strip()
                if page_text:
                    sections.append({"label": f"Page {page_index}", "text": page_text})
    except Exception:
        sections = []

    if sections:
        return sections

    try:
        convert_kwargs = {"dpi": 150}
        if POPLLER_PATH:
            convert_kwargs["poppler_path"] = POPLLER_PATH
        images = convert_from_bytes(pdf_bytes, first_page=1, last_page=max_pages, **convert_kwargs)
    except Exception:
        return []

    for page_index, image in enumerate(images, start=1):
        try:
            ocr_text = (pytesseract.image_to_string(image) or "").strip()
            if ocr_text:
                sections.append({"label": f"Page {page_index} (OCR)", "text": ocr_text})
        finally:
            try:
                image.close()
            except Exception:
                pass

    return sections


def extract_image_sections(image_bytes: bytes) -> List[Dict]:
    try:
        image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    except Exception:
        return []

    try:
        text = (pytesseract.image_to_string(image) or "").strip()
        if text:
            return [{"label": "Image OCR", "text": text}]
    except Exception:
        return []
    finally:
        try:
            image.close()
        except Exception:
            pass

    return []


def build_attachment_context(attachments: List[Dict], query_text: str = "", mode: str = "generic") -> str:
    attachment_blocks: List[str] = []
    query_tokens = tokenize_for_retrieval(query_text)
    max_sections = 6 if mode in {"multidoc", "compare"} else 3
    max_chunk_chars = 1400 if mode in {"multidoc", "compare"} else 1000

    for attachment in attachments or []:
        if not isinstance(attachment, dict):
            continue

        name = str(attachment.get("name") or "attachment")
        content = attachment.get("content") or attachment.get("raw") or ""
        if not isinstance(content, str) or not content.startswith("data:"):
            continue

        mime_type = str(attachment.get("type") or "").lower()
        file_bytes, data_mime_type = decode_data_url(content)
        mime_type = mime_type or data_mime_type
        lower_name = name.lower()

        if mime_type.startswith("image/") or lower_name.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff")):
            sections = extract_image_sections(file_bytes)
        elif mime_type == "application/pdf" or lower_name.endswith(".pdf"):
            sections = extract_pdf_sections(file_bytes)
        elif lower_name.endswith((".xlsx", ".xlsm", ".xls")) or mime_type in {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }:
            sections = extract_spreadsheet_sections(file_bytes, name)
        else:
            continue

        candidate_chunks: List[Dict] = []
        for section in sections:
            for chunk_index, chunk_text in enumerate(chunk_text_for_retrieval(section["text"], max_chars=max_chunk_chars), start=1):
                candidate_chunks.append({
                    "label": f"{section['label']} • Chunk {chunk_index}" if len(section["text"]) > max_chunk_chars else section["label"],
                    "text": chunk_text,
                    "score": score_chunk(query_tokens, chunk_text),
                })

        if not candidate_chunks:
            continue

        if query_tokens:
            candidate_chunks.sort(key=lambda item: (-item["score"], len(item["text"])))
        selected_chunks = candidate_chunks[:max_sections]

        formatted_chunks = []
        for chunk in selected_chunks:
            formatted_chunks.append(f"=== {name} | {chunk['label']} ===\n{mask_pii_text(chunk['text'])}")
        attachment_blocks.append("\n\n".join(formatted_chunks))

    if not attachment_blocks:
        return ""

    return "\n\n--- ATTACHMENT INDEX ---\n" + "\n\n".join(attachment_blocks)


def extract_text_from_image_bytes(image_bytes: bytes) -> str:
    try:
        image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    except Exception:
        return ""

    try:
        return (pytesseract.image_to_string(image) or "").strip()
    except Exception:
        return ""
    finally:
        try:
            image.close()
        except Exception:
            pass


def extract_text_from_pdf_bytes(pdf_bytes: bytes, max_pages: int = 3) -> str:
    text_parts: List[str] = []

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages[:max_pages]:
                page_text = (page.extract_text() or "").strip()
                if page_text:
                    text_parts.append(page_text)
    except Exception:
        text_parts = []

    combined_text = "\n\n".join(text_parts).strip()
    if combined_text:
        return combined_text

    try:
        convert_kwargs = {"dpi": 150}
        if POPLLER_PATH:
            convert_kwargs["poppler_path"] = POPLLER_PATH
        images = convert_from_bytes(pdf_bytes, first_page=1, last_page=max_pages, **convert_kwargs)
    except Exception:
        return ""

    ocr_parts: List[str] = []
    for image in images:
        try:
            ocr_text = (pytesseract.image_to_string(image) or "").strip()
            if ocr_text:
                ocr_parts.append(ocr_text)
        finally:
            try:
                image.close()
            except Exception:
                pass

    return "\n\n".join(ocr_parts).strip()


def call_azure_openai_chat(messages: List[Dict], system: str, model: Optional[str] = None, max_tokens: int = 2000) -> str:
    api_key = os.getenv("AZURE_OPENAI_API_KEY")
    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
    deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME")
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-01")

    if not api_key or not endpoint or not deployment_name:
        raise HTTPException(
            status_code=500,
            detail="Azure OpenAI credentials are missing. Check AZURE_OPENAI_API_KEY, AZURE_OPENAI_ENDPOINT, and AZURE_OPENAI_DEPLOYMENT_NAME in .env."
        )

    sanitized_messages = sanitize_chat_messages(messages)
    url = endpoint.rstrip("/") + f"/openai/deployments/{deployment_name}/chat/completions?api-version={api_version}"

    payload = {
        "messages": ([{"role": "system", "content": mask_pii_text(system)}] if system else []) + sanitized_messages,
        "max_tokens": max_tokens,
        "temperature": 0.2,
    }
    if model:
        payload["model"] = model

    request = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "api-key": api_key,
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            response_data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode("utf-8", errors="ignore") if hasattr(e, "read") else ""
        raise HTTPException(
            status_code=502,
            detail=error_body or f"Azure OpenAI request failed with status {e.code}"
        )
    except urllib.error.URLError as e:
        raise HTTPException(status_code=502, detail=f"Azure OpenAI request failed: {e.reason}")

    choices = response_data.get("choices", [])
    if not choices:
        raise HTTPException(status_code=502, detail="Azure OpenAI returned no choices")

    return choices[0].get("message", {}).get("content", "")


def translate_text_batch_with_azure(texts: List[str], source_lang: str, target_lang: str) -> List[str]:
    """Fallback translation using Azure OpenAI when local model is unavailable."""
    translated = []
    system_prompt = (
        f"You are a professional translation engine. Translate from {source_lang} to {target_lang}. "
        "Preserve meaning and structure. Return only translated text with no commentary."
    )

    for text in texts:
        messages = [{"role": "user", "content": text}]
        output = call_azure_openai_chat(messages, system_prompt, model="gpt-4o", max_tokens=1200)
        translated.append(unicodedata.normalize("NFC", output.strip()))

    return translated

# ============================================================================
# TRANSLATION RULES - IMPORTED FROM transliteration_rules.py
# ============================================================================
# All transliteration and word preservation rules are now in transliteration_rules.py
# To modify rules:
#   1. Edit TRANSLITERATE_WORDS dictionary (with language-specific versions)
#   2. Edit KEEP_ORIGINAL_WORDS set
#   3. Add/remove languages in LANGUAGE_CODE_TO_DICT_KEY mapping
#
# Example of adding a new word with multi-language support:
#   "new_word": {
#       "hindi": "हिंदी_संस्करण",
#       "tamil": "தமிழ்_பதிப்பு",
#       "telugu": "తెలుగు_సంస్కరణ",
#       ...
#   }
# ============================================================================
# FORM TRANSLATION PIPELINE - SEPARATE FROM REGULAR PDF TRANSLATION
# ============================================================================
# ============================================================================
# OPTIMIZED FORM TRANSLATION PIPELINE
# ============================================================================

import gc
from concurrent.futures import TimeoutError
import signal

# Add timeout decorator
class TimeoutException(Exception):
    pass

def timeout_handler(signum, frame):
    raise TimeoutException("Operation timed out")


def find_poppler_path() -> Optional[str]:
    """Find Poppler bin directory, preferring workspace-local installations."""
    cwd = os.getcwd()
    candidates = [
        os.path.join(cwd, 'poppler', 'Library', 'bin'),
        os.path.join(cwd, 'poppler', 'bin'),
    ]
    candidates.extend(glob.glob(os.path.join(cwd, 'poppler-*', 'Library', 'bin')))
    candidates.extend(glob.glob(os.path.join(cwd, 'poppler-*', 'bin')))

    for path in candidates:
        if os.path.isdir(path) and os.path.exists(os.path.join(path, 'pdfinfo.exe')):
            print(f"Using Poppler from: {path}")
            return path

    print("Poppler binary not found in workspace-local folders.")
    return None

def extract_form_structure_optimized(pdf_path: str, max_pages: int = 50) -> Dict:
    """
    OPTIMIZED: Extract form structure with memory limits and early exits
    """
    doc = fitz.open(pdf_path)
    
    # Safety check
    if len(doc) > max_pages:
        doc.close()
        raise HTTPException(
            status_code=400, 
            detail=f"Form too large ({len(doc)} pages). Max {max_pages} pages allowed."
        )
    
    form_structure = {
        'pages': [],
        'total_fields': 0,
        'field_types': {},
        'is_scanned': False
    }
    
    try:
        for page_num in range(len(doc)):
            page = doc[page_num]
            page_data = {
                'page_number': page_num,
                'fields': [],
                'text_blocks': []  # Simplified - don't distinguish labels/static
            }
            
            # Extract form fields (FAST)
            widgets = page.widgets()
            field_rects = []
            SAFETY_MARGIN = 5
            for widget in widgets:
                try:
                    field_info = {
                        'name': widget.field_name or f"field_{page_num}_{len(page_data['fields'])}",
                        'type': widget.field_type_string or "Unknown",
                        'rect': list(widget.rect),
                    }
                    page_data['fields'].append(field_info)
                    original_rect = fitz.Rect(widget.rect)
                    expanded_rect = fitz.Rect(
                        original_rect.x0 - SAFETY_MARGIN,
                        original_rect.y0 - SAFETY_MARGIN,
                        original_rect.x1 + SAFETY_MARGIN,
                        original_rect.y1 + SAFETY_MARGIN
                    )
                    field_rects.append(expanded_rect)
                    
                    # Count field types
                    field_type = field_info['type']
                    form_structure['field_types'][field_type] = form_structure['field_types'].get(field_type, 0) + 1
                except Exception as e:
                    print(f"Warning: Skipped problematic field on page {page_num}: {e}")
                    continue
            
            form_structure['total_fields'] += len(page_data['fields'])
            
            # Extract text blocks (OPTIMIZED - no complex distance calculations)
            try:
                blocks = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]
                
                for block in blocks:
                    if block.get("type") != 0:  # Skip non-text
                        continue
                    
                    for line in block.get("lines", []):
                        line_text = ""
                        line_bbox = line.get("bbox")
                        font_size = 12
                        font_color = (0, 0, 0)
                        is_bold = False
                        
                        for span in line.get("spans", []):
                            line_text += span.get("text", "")
                            font_size = span.get('size', 12)
                            
                            # Check bold
                            font_flags = span.get('flags', 0)
                            is_bold = bool(font_flags & 2**4)
                            
                            # Color
                            color = span.get('color', 0)
                            if isinstance(color, int):
                                r = ((color >> 16) & 0xFF) / 255.0
                                g = ((color >> 8) & 0xFF) / 255.0
                                b = (color & 0xFF) / 255.0
                                font_color = (r, g, b)
                        
                        if not line_text.strip():
                            continue
                        if should_skip_text(line_text):
                            continue
                        line_rect = fitz.Rect(line_bbox)
                        
                        # OPTIMIZED: Simple overlap check instead of distance calculation
                        is_inside_field = False
                        for field_rect in field_rects:
                            if line_rect.intersects(field_rect):
                                is_inside_field = True
                                break
                        
                        # Only translate text NOT inside form fields
                        if not is_inside_field:
                            # Additional check: ensure text doesn't overlap vertically
                            text_is_safe = True
                            for field_rect in field_rects:
                                # Check if text is in same vertical region as field
                                if (line_rect.y0 < field_rect.y1 and line_rect.y1 > field_rect.y0):
                                    # If horizontally close, skip it
                                    horizontal_distance = min(
                                        abs(line_rect.x1 - field_rect.x0),
                                        abs(field_rect.x1 - line_rect.x0)
                                    )
                                    if horizontal_distance < 20:  # Less than 20 pixels away
                                        text_is_safe = False
                                        break
                            
                            if text_is_safe:
                                page_data['text_blocks'].append({
                                    'text': line_text.strip(),
                                    'bbox': line_bbox,
                                    'font_size': font_size,
                                    'font_color': font_color,
                                    'is_bold': is_bold
                                })
            
            except Exception as e:
                print(f"Warning: Text extraction failed on page {page_num}: {e}")
            
            form_structure['pages'].append(page_data)
            
            # Force garbage collection every 10 pages
            if page_num % 10 == 0:
                gc.collect()
        
    finally:
        doc.close()
        gc.collect()
    
    return form_structure


def translate_form_structure_batched(form_structure: Dict, source_lang: str, target_lang: str, batch_size: int = 20) -> Dict:
    """
    OPTIMIZED: Batch translation with memory limits
    """
    translated_structure = form_structure.copy()
    translated_structure['pages'] = []
    
    for page_idx, page_data in enumerate(form_structure['pages']):
        print(f"Translating page {page_idx + 1}/{len(form_structure['pages'])}...")
        
        translated_page = page_data.copy()
        translated_page['text_blocks'] = []
        
        text_blocks = page_data.get('text_blocks', [])
        
        if not text_blocks:
            translated_structure['pages'].append(translated_page)
            continue
        
        # Process in batches to avoid memory issues
        total_blocks = len(text_blocks)
        
        for i in range(0, total_blocks, batch_size):
            batch = text_blocks[i:i + batch_size]
            texts = [block['text'] for block in batch]
            
            try:
                # Translate batch with timeout
                translations = translate_text_batch(texts, source_lang, target_lang)
                
                for original_block, translated_text in zip(batch, translations):
                    translated_block = original_block.copy()
                    translated_block['original'] = original_block['text']
                    translated_block['translated'] = translated_text
                    translated_page['text_blocks'].append(translated_block)
                
                print(f"  Batch {i//batch_size + 1}: Translated {len(batch)} blocks")
                
            except Exception as e:
                print(f"  Warning: Batch {i//batch_size + 1} failed: {e}")
                # Add originals as fallback
                for original_block in batch:
                    fallback_block = original_block.copy()
                    fallback_block['original'] = original_block['text']
                    fallback_block['translated'] = original_block['text']  # Keep original
                    translated_page['text_blocks'].append(fallback_block)
            
            # Clear memory
            gc.collect()
        
        translated_structure['pages'].append(translated_page)
    
    return translated_structure


def create_translated_form_pdf_optimized(
    original_pdf_path: str,
    translated_structure: Dict,
    output_path: str,
    target_lang: str
):
    """
    OPTIMIZED: Create translated form with better error handling
    """
    font_path = get_font_for_language(target_lang)
    
    if not font_path or not os.path.exists(font_path):
        raise Exception(f"Font not found for {target_lang}")
    
    # Find bold font
    bold_font_path = None
    font_dir = os.path.dirname(font_path)
    for bold_name in ['Bold', 'bold', 'BOLD']:
        potential_bold = font_path.replace('.ttf', f'-{bold_name}.ttf')
        if os.path.exists(potential_bold):
            bold_font_path = potential_bold
            break
    
    print(f"Using font: {font_path}")
    
    # Open original PDF
    doc = fitz.open(original_pdf_path)
    
    try:
        font_name = "unicode_font"
        bold_font_name = "unicode_font_bold"
        
        for page_data in translated_structure['pages']:
            page_num = page_data['page_number']
            
            if page_num >= len(doc):
                break
            
            page = doc[page_num]
            text_blocks = page_data.get('text_blocks', [])
            
            if not text_blocks:
                continue
            
            print(f"Page {page_num + 1}: Processing {len(text_blocks)} blocks...")
            
            # Redact originals
            for block in text_blocks:
                if 'original' not in block:
                    continue
                
                try:
                    bbox = block['bbox']
                    rect = fitz.Rect(bbox)
                    page.add_redact_annot(rect)
                except Exception as e:
                    print(f"  Warning: Redaction failed for block: {e}")
            
            # Apply redactions
            try:
                page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
            except Exception as e:
                print(f"  Warning: Apply redactions failed: {e}")
            
            # Insert translations
            for block in text_blocks:
                if 'translated' not in block:
                    continue
                
                try:
                    bbox = block['bbox']
                    rect = fitz.Rect(bbox)
                    translated_text = unicodedata.normalize("NFC", block['translated'])
                    
                    original_font_size = block.get('font_size', 12)
                    font_color = block.get('font_color', (0, 0, 0))
                    is_bold = block.get('is_bold', False)
                    
                    current_font_path = bold_font_path if (is_bold and bold_font_path) else font_path
                    current_font_name = bold_font_name if (is_bold and bold_font_path) else font_name
                    
                    # Try inserting with original size
                    rc = page.insert_textbox(
                        rect,
                        translated_text,
                        fontname=current_font_name,
                        fontfile=current_font_path,
                        fontsize=original_font_size,
                        color=font_color,
                        align=fitz.TEXT_ALIGN_LEFT
                    )
                    
                    # Quick size reduction if needed (fewer iterations)
                    if rc < 0:
                        for reduction in [0.85, 0.70, 0.55]:
                            new_size = original_font_size * reduction
                            rc = page.insert_textbox(
                                rect, translated_text,
                                fontname=current_font_name,
                                fontfile=current_font_path,
                                fontsize=new_size,
                                color=font_color,
                                align=fitz.TEXT_ALIGN_LEFT
                            )
                            if rc >= 0:
                                break
                
                except Exception as e:
                    print(f"  Warning: Text insertion failed: {e}")
            
            # Clear memory every page
            gc.collect()
        
        # Save
        print("Saving translated PDF...")
        doc.save(output_path, garbage=4, deflate=True)
    
    finally:
        doc.close()
        gc.collect()


def process_scanned_form_optimized(pdf_path: str, dpi: int = 150, max_pages: int = 20) -> Dict:
    """
    OPTIMIZED: Process scanned form with lower DPI and page limits
    """
    print(f"Processing scanned form (DPI: {dpi}, max pages: {max_pages})...")
    
    # Check page count first
    doc = fitz.open(pdf_path)
    page_count = len(doc)
    doc.close()
    
    if page_count > max_pages:
        raise HTTPException(
            status_code=400,
            detail=f"Scanned form too large ({page_count} pages). Max {max_pages} pages. Use lower DPI or split document."
        )
    
    pdf_bytes = open(pdf_path, 'rb').read()
    
    try:
        # Use lower DPI for memory efficiency
        convert_kwargs = {'dpi': dpi}
        if POPLLER_PATH:
            convert_kwargs['poppler_path'] = POPLLER_PATH
        images = convert_from_bytes(pdf_bytes, **convert_kwargs)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF to image conversion failed: {e}")
    
    form_structure = {
        'pages': [],
        'total_fields': 0,
        'field_types': {},
        'is_scanned': True
    }
    
    for page_idx, image in enumerate(images):
        print(f"OCR on page {page_idx + 1}/{len(images)}...")
        
        page_data = {
            'page_number': page_idx,
            'fields': [],
            'text_blocks': []
        }
        
        try:
            # OCR with timeout protection
            ocr_data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
            
            # Group into lines
            lines = {}
            n_boxes = len(ocr_data['text'])
            
            for i in range(n_boxes):
                text = ocr_data['text'][i].strip()
                conf = int(ocr_data['conf'][i]) if ocr_data['conf'][i] != '-1' else 0
                
                if text and conf > 40:  # Higher confidence threshold
                    line_num = ocr_data['line_num'][i]
                    if line_num not in lines:
                        lines[line_num] = []
                    lines[line_num].append({
                        'text': text,
                        'x': ocr_data['left'][i],
                        'y': ocr_data['top'][i],
                        'width': ocr_data['width'][i],
                        'height': ocr_data['height'][i]
                    })
            
            # Process lines
            for line_words in lines.values():
                if not line_words:
                    continue
                
                combined_text = ' '.join([w['text'] for w in line_words])
                x = min([w['x'] for w in line_words])
                y = min([w['y'] for w in line_words])
                max_x = max([w['x'] + w['width'] for w in line_words])
                max_y = max([w['y'] + w['height'] for w in line_words])
                
                page_data['text_blocks'].append({
                    'text': combined_text,
                    'bbox': [x, y, max_x, max_y],
                    'font_size': max(8, min(max_y - y, 24)),
                    'font_color': (0, 0, 0),
                    'is_bold': False
                })
        
        except Exception as e:
            print(f"Warning: OCR failed on page {page_idx + 1}: {e}")
        
        form_structure['pages'].append(page_data)
        
        # Clear memory
        del image
        gc.collect()
    
    return form_structure
def get_language_token_id(tokenizer, lang_code: str) -> int:
    """Get the token ID for a language code in NLLB tokenizer"""
    try:
        return tokenizer.convert_tokens_to_ids(lang_code)
    except:
        try:
            return tokenizer.lang_code_to_id[lang_code]
        except:
            tokens = tokenizer.encode(lang_code, add_special_tokens=False)
            return tokens[0] if tokens else tokenizer.pad_token_id

def find_fonts():
    """Find Noto fonts on the system"""
    fonts_dict = {}
    cwd = os.getcwd()
    script_dirs = []

    try:
        for entry in os.scandir(cwd):
            if entry.is_dir() and entry.name.lower().startswith(('noto_', 'tiro_', 'hind_', 'fonts', 'font')):
                script_dirs.append(entry.path)
    except Exception:
        pass
    
    search_paths = [
        cwd,
        'C:\\Windows\\Fonts',
        os.path.expanduser('~\\AppData\\Local\\Microsoft\\Windows\\Fonts'),
        '/usr/share/fonts/truetype/noto',
        '/usr/share/fonts/google-noto',
        '/usr/share/fonts/truetype',
        '/System/Library/Fonts',
        '/Library/Fonts',
        './fonts',
        '../fonts',
        os.path.join(os.getcwd(), 'fonts'),
    ]
    search_paths.extend(script_dirs)
    
    print("Searching for fonts...")
    for path in search_paths:
        path = os.path.expanduser(path)
        if os.path.exists(path):
            patterns = [
                '**/Noto*.ttf', '**/noto*.ttf', 'Noto*.ttf', 'noto*.ttf',
                '**/Hind*.ttf', '**/hind*.ttf', '**/Tiro*.ttf', '**/tiro*.ttf',
                '**/*.otf'
            ]
            for pattern in patterns:
                found_fonts = glob.glob(os.path.join(path, pattern), recursive=True)
                for font in found_fonts:
                    font_name = os.path.basename(font).lower()
                    if ('devanagari' in font_name or 'tirodevanagari' in font_name) and 'devanagari' not in fonts_dict:
                        fonts_dict['devanagari'] = font
                        print(f"  ✓ Devanagari: {font}")
                    elif ('bengali' in font_name or 'siliguri' in font_name) and 'bengali' not in fonts_dict:
                        fonts_dict['bengali'] = font
                        print(f"  ✓ Bengali: {font}")
                    elif 'tamil' in font_name and 'tamil' not in fonts_dict:
                        fonts_dict['tamil'] = font
                        print(f"  ✓ Tamil: {font}")
                    elif 'telugu' in font_name and 'telugu' not in fonts_dict:
                        fonts_dict['telugu'] = font
                        print(f"  ✓ Telugu: {font}")
                    elif 'gujarati' in font_name and 'gujarati' not in fonts_dict:
                        fonts_dict['gujarati'] = font
                        print(f"  ✓ Gujarati: {font}")
                    elif 'kannada' in font_name and 'kannada' not in fonts_dict:
                        fonts_dict['kannada'] = font
                        print(f"  ✓ Kannada: {font}")
                    elif 'malayalam' in font_name and 'malayalam' not in fonts_dict:
                        fonts_dict['malayalam'] = font
                        print(f"  ✓ Malayalam: {font}")
                    elif ('oriya' in font_name or 'odia' in font_name) and 'oriya' not in fonts_dict:
                        fonts_dict['oriya'] = font
                        print(f"  ✓ Oriya/Odia: {font}")
                    elif 'notosans' in font_name and 'regular' in font_name and 'default' not in fonts_dict:
                        fonts_dict['default'] = font
                        print(f"  ✓ Default: {font}")
                    elif ('hind' in font_name or 'tiro' in font_name) and 'default' not in fonts_dict:
                        fonts_dict['default'] = font
                        print(f"  ✓ Default: {font}")
    
    return fonts_dict

def load_models():
    """Load translation models on startup"""
    global text_model, text_tokenizer, device, FONTS, POPLLER_PATH

    # Avoid HF Xet streaming issues on some Windows setups.
    os.environ.setdefault("HF_HUB_DISABLE_XET", "1")
    
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Using device: {device}")
    
    POPLLER_PATH = find_poppler_path()
    FONTS = find_fonts()
    
    print("\nLoading text translation model (NLLB-200)...")
    text_model_name = "facebook/nllb-200-distilled-600M"
    text_tokenizer = AutoTokenizer.from_pretrained(text_model_name)
    text_model = AutoModelForSeq2SeqLM.from_pretrained(
        text_model_name,
        torch_dtype=torch.float16 if torch.cuda.is_available() else torch.float32
    ).to(device)
    text_model.eval()
    print("Text translation model loaded!")

def get_language_code(lang_name: str) -> str:
    """Convert language name to NLLB language code"""
    lang_map = {
        'english': 'eng_Latn', 'spanish': 'spa_Latn', 'french': 'fra_Latn',
        'german': 'deu_Latn', 'italian': 'ita_Latn', 'portuguese': 'por_Latn',
        'russian': 'rus_Cyrl', 'hindi': 'hin_Deva', 'bengali': 'ben_Beng',
        'bangla': 'ben_Beng', 'tamil': 'tam_Taml', 'telugu': 'tel_Telu',
        'marathi': 'mar_Deva', 'gujarati': 'guj_Gujr', 'kannada': 'kan_Knda',
        'malayalam': 'mal_Mlym', 'punjabi': 'pan_Guru', 'oriya': 'ory_Orya',
        'odia': 'ory_Orya', 'assamese': 'asm_Beng', 'urdu': 'urd_Arab',
    }
    return lang_map.get(lang_name.lower().strip(), 'eng_Latn')

def get_font_for_language(target_lang: str) -> str:
    """Get appropriate font for target language"""
    lang_lower = target_lang.lower()
    
    if any(x in lang_lower for x in ['hindi', 'marathi', 'sanskrit', 'nepali']):
        return FONTS.get('devanagari', FONTS.get('default'))
    elif any(x in lang_lower for x in ['bengali', 'bangla', 'assamese']):
        return FONTS.get('bengali', FONTS.get('default'))
    elif 'tamil' in lang_lower:
        return FONTS.get('tamil', FONTS.get('default'))
    elif 'telugu' in lang_lower:
        return FONTS.get('telugu', FONTS.get('default'))
    elif 'gujarati' in lang_lower:
        return FONTS.get('gujarati', FONTS.get('default'))
    elif 'kannada' in lang_lower:
        return FONTS.get('kannada', FONTS.get('default'))
    elif 'malayalam' in lang_lower:
        return FONTS.get('malayalam', FONTS.get('default'))
    elif any(x in lang_lower for x in ['odia', 'oriya']):
        return FONTS.get('oriya', FONTS.get('default'))
    return FONTS.get('default')

def translate_text_batch(texts: List[str], source_lang: str, target_lang: str) -> List[str]:
    """Translate multiple texts using NLLB model with custom rules"""
    if not texts:
        return []

    if text_model is None or text_tokenizer is None:
        return translate_text_batch_with_azure(texts, source_lang, target_lang)

    texts = [mask_pii_text(text) for text in texts]
    
    src_code = get_language_code(source_lang)
    tgt_code = get_language_code(target_lang)
    
    text_tokenizer.src_lang = src_code
    inputs = text_tokenizer(texts, return_tensors="pt", padding=True, truncation=True, max_length=512).to(device)
    forced_bos_token_id = get_language_token_id(text_tokenizer, tgt_code)
    
    with torch.no_grad():
        translated_tokens = text_model.generate(
            **inputs, forced_bos_token_id=forced_bos_token_id,
            max_length=512, num_beams=5, early_stopping=True
        )
    
    translations = text_tokenizer.batch_decode(translated_tokens, skip_special_tokens=True)
    
    # Apply custom rules to each translation with target language code for language-specific transliterations
    processed_translations = []
    for original, translated in zip(texts, translations):
        processed = apply_translation_rules(original, translated, tgt_code)
        processed_translations.append(unicodedata.normalize("NFC", processed))
    
    return processed_translations

# apply_translation_rules is now imported from transliteration_rules.py
# It includes full multi-language support

def is_scanned_pdf(pdf_path: str) -> bool:
    """Determine if PDF is scanned or text-based"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages_to_check = min(3, len(pdf.pages))
            total_chars = sum(len(pdf.pages[i].extract_text() or '') for i in range(pages_to_check))
            return total_chars < 100
    except:
        return True

def extract_text_blocks_with_pymupdf(pdf_path: str) -> List[List[Dict]]:
    """Extract text blocks with positions and font properties using PyMuPDF"""
    doc = fitz.open(pdf_path)
    all_pages_blocks = []
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        blocks = page.get_text("dict")["blocks"]
        
        page_blocks = []
        for block in blocks:
            if block.get("type") == 0:  # Text block
                for line in block.get("lines", []):
                    line_text = ""
                    line_bbox = line.get("bbox")
                    font_size = 12
                    font_name = "helv"
                    font_color = (0, 0, 0)  # Default black
                    
                    for span in line.get("spans", []):
                        line_text += span.get("text", "")
                        font_size = span.get('size', 12)
                        font_name = span.get('font', 'helv')
                        
                        # Extract color (RGB values 0-1)
                        color = span.get('color', 0)
                        if isinstance(color, int):
                            # Convert integer color to RGB
                            r = ((color >> 16) & 0xFF) / 255.0
                            g = ((color >> 8) & 0xFF) / 255.0
                            b = (color & 0xFF) / 255.0
                            font_color = (r, g, b)
                    
                    if line_text.strip():
                        page_blocks.append({
                            'text': line_text.strip(),
                            'bbox': line_bbox,
                            'font_size': font_size,
                            'font_name': font_name,
                            'font_color': font_color  # Store original color
                        })
        
        all_pages_blocks.append(page_blocks)
    
    doc.close()
    return all_pages_blocks
def should_skip_text(text: str) -> bool:
    """
    Skip single letters used as bullet points (a, b, c, i, ii, iii, etc.)
    """
    text_stripped = text.strip()
    
    # Skip single letters followed by punctuation (a., b), etc.)
    if len(text_stripped) <= 3:
        # Single letter bullet points
        if text_stripped in ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k']:
            return True
        if text_stripped in ['a.', 'b.', 'c.', 'd.', 'e.', 'f.', 'g.', 'h.']:
            return True
        if text_stripped in ['a)', 'b)', 'c)', 'd)', 'e)', 'f)', 'g)', 'h)']:
            return True
        # Roman numerals
        if text_stripped.lower() in ['i', 'ii', 'iii', 'iv', 'v', 'vi', 'vii', 'viii', 'ix', 'x']:
            return True
        if text_stripped.lower() in ['i.', 'ii.', 'iii.', 'iv.', 'v.']:
            return True
    
    return False
def create_translated_pdf_proper(
    original_pdf_path: str,
    text_blocks_per_page: List[List[Dict]],
    output_path: str,
    target_lang: str
):
    """
    Proper PDF text replacement with font property preservation:
    1. Use redaction to remove original text
    2. Register Unicode font in PDF
    3. Insert translated text with ORIGINAL font size, color, and bold
    """
    font_path = get_font_for_language(target_lang)
    
    if not font_path or not os.path.exists(font_path):
        raise Exception(f"Font not found for {target_lang}. Please install Noto fonts in ./fonts/ or C:\\Windows\\Fonts")
    
    # Also try to find bold font
    bold_font_path = None
    font_dir = os.path.dirname(font_path)
    font_basename = os.path.basename(font_path).lower()
    
    # Look for bold variant
    for bold_name in ['bold', 'Bold', 'BOLD']:
        potential_bold = font_path.replace('.ttf', f'-{bold_name}.ttf')
        if os.path.exists(potential_bold):
            bold_font_path = potential_bold
            break
    
    # If not found, try searching in same directory
    if not bold_font_path and os.path.exists(font_dir):
        for file in os.listdir(font_dir):
            if 'bold' in file.lower() and font_basename.split('-')[0] in file.lower():
                bold_font_path = os.path.join(font_dir, file)
                break
    
    print(f"Using regular font: {font_path}")
    if bold_font_path:
        print(f"Using bold font: {bold_font_path}")
    else:
        print("Bold font not found - will use regular font with increased weight")
    
    # Open original PDF
    doc = fitz.open(original_pdf_path)
    
    # Register fonts
    font_name = "unicode_font"
    bold_font_name = "unicode_font_bold"
    
    try:
        with open(font_path, 'rb') as f:
            font_data = f.read()
        if bold_font_path:
            with open(bold_font_path, 'rb') as f:
                bold_font_data = f.read()
    except Exception as e:
        raise Exception(f"Could not read font file: {e}")
    
    for page_num, text_blocks in enumerate(text_blocks_per_page):
        if page_num >= len(doc):
            break
        
        page = doc[page_num]
        
        # First pass: Redact (remove) original text
        for block in text_blocks:
            if not block.get('original'):
                continue
            
            bbox = block['bbox']
            rect = fitz.Rect(bbox)
            
            # Add redaction annotation
            page.add_redact_annot(rect)
        
        # Apply all redactions at once
        page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
        
        # Second pass: Insert translated text with PRESERVED font properties
        for block in text_blocks:
            if not block.get('translated'):
                continue
            
            bbox = block['bbox']
            rect = fitz.Rect(bbox)
            translated_text = unicodedata.normalize("NFC", block['translated'])
            
            # USE ORIGINAL FONT PROPERTIES
            original_font_size = block.get('font_size', 12)
            font_color = block.get('font_color', (0, 0, 0))
            original_font_name = block.get('font_name', '').lower()
            
            # Detect if original was bold
            is_bold = any(keyword in original_font_name for keyword in ['bold', 'heavy', 'black', 'demi'])
            
            # PRESERVE LARGER SIZES - Don't reduce large text
            # Only scale down if text doesn't fit, but try to keep size
            font_size = original_font_size
            
            # Use bold font if available and text was bold
            current_font_path = bold_font_path if (is_bold and bold_font_path) else font_path
            current_font_name = bold_font_name if (is_bold and bold_font_path) else font_name
            
            # Make color darker/bolder for bold text if no bold font available
            if is_bold and not bold_font_path:
                # Darken the color by 20% to simulate bold
                font_color = tuple(max(0, c * 0.8) for c in font_color)
            
            print(f"  Inserting: '{translated_text[:30]}...' (size: {font_size:.1f}, bold: {is_bold}, color: RGB{font_color})")
            
            # Insert text with the custom font and ORIGINAL properties
            rc = page.insert_textbox(
                rect,
                translated_text,
                fontname=current_font_name,
                fontfile=current_font_path,
                fontsize=font_size,      # PRESERVED from original
                color=font_color,         # PRESERVED from original
                align=fitz.TEXT_ALIGN_LEFT
            )
            
            # If text didn't fit, try smaller sizes BUT maintain relative proportions
            if rc < 0:
                # Try less aggressive reductions to preserve size hierarchy
                for reduction in [0.95, 0.90, 0.85, 0.80, 0.75, 0.70]:
                    new_size = original_font_size * reduction
                    rc = page.insert_textbox(
                        rect,
                        translated_text,
                        fontname=current_font_name,
                        fontfile=current_font_path,
                        fontsize=new_size,
                        color=font_color,
                        align=fitz.TEXT_ALIGN_LEFT
                    )
                    if rc >= 0:
                        print(f"    Reduced font to {new_size:.1f} (from {original_font_size:.1f}) to fit")
                        break
                
                # Last resort: try multi-line
                if rc < 0:
                    for reduction in [0.65, 0.60, 0.55]:
                        new_size = original_font_size * reduction
                        rc = page.insert_textbox(
                            rect,
                            translated_text,
                            fontname=current_font_name,
                            fontfile=current_font_path,
                            fontsize=new_size,
                            color=font_color,
                            align=fitz.TEXT_ALIGN_LEFT
                        )
                        if rc >= 0:
                            print(f"    Significantly reduced to {new_size:.1f} (multi-line)")
                            break
    
    # Save with embedded fonts
    doc.save(output_path, garbage=4, deflate=True)
    doc.close()

def ocr_pdf_page(image: Image.Image) -> List[Dict]:
    """Perform OCR on a PDF page image"""
    ocr_data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
    text_blocks = []
    lines = {}
    
    n_boxes = len(ocr_data['text'])
    for i in range(n_boxes):
        text = ocr_data['text'][i].strip()
        conf = int(ocr_data['conf'][i]) if ocr_data['conf'][i] != '-1' else 0
        
        if text and conf > 30:
            line_num = ocr_data['line_num'][i]
            if line_num not in lines:
                lines[line_num] = []
            lines[line_num].append({
                'text': text, 'x': ocr_data['left'][i], 'y': ocr_data['top'][i],
                'width': ocr_data['width'][i], 'height': ocr_data['height'][i]
            })
    
    for line_words in lines.values():
        if not line_words:
            continue
        
        combined_text = ' '.join([w['text'] for w in line_words])
        x = min([w['x'] for w in line_words])
        y = min([w['y'] for w in line_words])
        max_x = max([w['x'] + w['width'] for w in line_words])
        max_y = max([w['y'] + w['height'] for w in line_words])
        
        text_blocks.append({
            'text': combined_text,
            'bbox': [x, y, max_x, max_y],
            'font_size': max(8, min(max_y - y, 24))
        })
    
    return text_blocks
@router.websocket("/ws/{client_id}")
async def websocket_endpoint(websocket: WebSocket, client_id: str):
    await manager.connect(websocket, client_id)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(client_id)
@router.post("/translate-pdf/")
async def translate_pdf(
    file: UploadFile = File(...),
    source_lang: str = Form(...),
    target_lang: str = Form(...),
    dpi: int = Form(200),
    client_id: str = Form(None)
):
    """Translate a PDF with proper text replacement (no boxes/images)"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
    
    async def log(msg: str):
        print(msg)
        if client_id:
            await manager.send_message(msg, client_id)
    
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_input:
            content = await file.read()
            tmp_input.write(content)
            input_path = tmp_input.name
        
        await log(f"Processing PDF: {file.filename}")
        await log(f"Translation: {source_lang} → {target_lang}")
        
        is_scanned = is_scanned_pdf(input_path)
        await log(f"PDF type: {'Scanned (OCR)' if is_scanned else 'Text-based'}")
        
        text_blocks_per_page = []
        
        if is_scanned:
            await log("Converting to images for OCR...")
            pdf_bytes = open(input_path, 'rb').read()
            convert_kwargs = {'dpi': dpi}
            if POPLLER_PATH:
                convert_kwargs['poppler_path'] = POPLLER_PATH
            images = convert_from_bytes(pdf_bytes, **convert_kwargs)
            
            for page_idx, image in enumerate(images):
                await log(f"Page {page_idx + 1}/{len(images)}: OCR...")
                text_blocks = ocr_pdf_page(image)
                
                if text_blocks:
                    texts = [mask_pii_text(b['text']) for b in text_blocks]
                    translations = translate_text_batch(texts, source_lang, target_lang)
                    
                    for block, masked_text, trans in zip(text_blocks, texts, translations):
                        block['original'] = masked_text
                        block['translated'] = trans
                
                text_blocks_per_page.append(text_blocks)
                await log(f"Page {page_idx + 1}/{len(images)}: Done ✓")
        else:
            await log("Extracting text with positions...")
            text_blocks_per_page = extract_text_blocks_with_pymupdf(input_path)
            
            total_pages = len(text_blocks_per_page)
            for page_idx, text_blocks in enumerate(text_blocks_per_page):
                if not text_blocks:
                    continue
                
                await log(f"Page {page_idx + 1}/{total_pages}: Translating...")
                
                texts = [mask_pii_text(b['text']) for b in text_blocks]
                translations = translate_text_batch(texts, source_lang, target_lang)
                
                for block, masked_text, trans in zip(text_blocks, texts, translations):
                    block['original'] = masked_text
                    block['translated'] = trans
                
                await log(f"Page {page_idx + 1}/{total_pages}: Done ✓")
        
        await log("Creating translated PDF...")
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_output:
            output_path = tmp_output.name
        
        create_translated_pdf_proper(input_path, text_blocks_per_page, output_path, target_lang)
        
        os.unlink(input_path)
        
        await log("Translation complete! ✓")
        return FileResponse(
            output_path,
            media_type='application/pdf',
            filename=f"translated_{file.filename}"
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        await log(f"Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Translation failed: {str(e)}")
@router.post("/translate-form/")
async def translate_form(
    file: UploadFile = File(...),
    source_lang: str = Form(...),
    target_lang: str = Form(...),
    dpi: int = Form(150),
    max_pages: int = Form(30),
    client_id: str = Form(None)
):
    """OPTIMIZED FORM PIPELINE with memory limits and timeouts"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
    
    async def log(msg: str):
        print(msg)
        if client_id:
            await manager.send_message(msg, client_id)
    
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_input:
            content = await file.read()
            
            file_size_mb = len(content) / (1024 * 1024)
            if file_size_mb > 50:
                raise HTTPException(
                    status_code=400,
                    detail=f"File too large ({file_size_mb:.1f}MB). Max 50MB for forms."
                )
            
            tmp_input.write(content)
            input_path = tmp_input.name
        
        await log(f"Processing form: {file.filename} ({file_size_mb:.1f}MB)")
        await log(f"Translation: {source_lang} → {target_lang}")
        
        is_scanned = is_scanned_pdf(input_path)
        await log(f"PDF type: {'Scanned (OCR)' if is_scanned else 'Digital form'}")
        
        if is_scanned:
            await log(f"Extracting scanned form (DPI: {dpi})...")
            form_structure = process_scanned_form_optimized(input_path, dpi, max_pages)
        else:
            await log(f"Extracting form structure...")
            form_structure = extract_form_structure_optimized(input_path, max_pages)
        
        total_pages = len(form_structure['pages'])
        await log(f"Form analysis: {form_structure['total_fields']} fields, {total_pages} pages")
        
        await log(f"Translating form content...")
        translated_structure = translate_form_structure_batched(
            form_structure, 
            source_lang, 
            target_lang,
            batch_size=15
        )
        
        # Add page-by-page logging in translate_form_structure_batched
        for page_idx in range(total_pages):
            await log(f"Page {page_idx + 1}/{total_pages}: Done ✓")
        
        await log(f"Creating translated PDF...")
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_output:
            output_path = tmp_output.name
        
        create_translated_form_pdf_optimized(
            input_path,
            translated_structure,
            output_path,
            target_lang
        )
        
        os.unlink(input_path)
        gc.collect()
        
        await log(f"Form translation complete! ✓")
        
        return FileResponse(
            output_path,
            media_type='application/pdf',
            filename=f"translated_form_{file.filename}"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        await log(f"Error: {str(e)}")
        
        try:
            if 'input_path' in locals():
                os.unlink(input_path)
        except:
            pass
        
        gc.collect()
        raise HTTPException(status_code=500, detail=f"Form translation failed: {str(e)}")

@router.get("/analyze-form/")
async def analyze_form(file: UploadFile = File(...)):
    """
    Analyze form structure without translating
    Useful for debugging and understanding form layout
    """
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_input:
            content = await file.read()
            tmp_input.write(content)
            input_path = tmp_input.name
        
        is_scanned = is_scanned_pdf(input_path)
        
        if is_scanned:
            form_structure = process_scanned_form_optimized(input_path)
        else:
            form_structure = extract_form_structure_optimized(input_path)
        
        os.unlink(input_path)
        
        # Simplify for JSON response
        simplified_structure = {
            'total_fields': form_structure['total_fields'],
            'field_types': form_structure['field_types'],
            'is_scanned': form_structure.get('is_scanned', False),
            'pages': []
        }
        
        for page_data in form_structure['pages']:
            simplified_structure['pages'].append({
                'page_number': page_data['page_number'],
                'field_count': len(page_data['fields']),
                'label_count': len(page_data['labels']),
                'static_text_count': len(page_data['static_text']),
                'sample_labels': [l['text'][:50] for l in page_data['labels'][:3]],
                'sample_fields': [
                    {
                        'name': f['name'],
                        'type': f['type']
                    } for f in page_data['fields'][:3]
                ]
            })
        
        return simplified_structure
        
    except Exception as e:
        import traceback
        traceback.print_exc()

@router.get("/languages")
async def list_languages():
    return {
        "indian_languages": ["Hindi", "Bengali", "Tamil", "Telugu", "Marathi", "Gujarati", "Kannada", "Malayalam"],
        "fonts_status": {lang: "✓" if lang in FONTS else "✗" for lang in ['devanagari', 'bengali', 'tamil', 'telugu']}
    }

@router.get("/translation-rules")
async def get_translation_rules():
    """Get current translation rules"""
    return {
        "transliterate_words": TRANSLITERATE_WORDS,
        "keep_original_words": list(KEEP_ORIGINAL_WORDS),
        "info": {
            "transliterate": "These words will be transliterated (e.g., 'google' → 'गूगल')",
            "keep_original": "These words will NOT be translated (stay in original language)"
        }
    }

@router.get("/test-rules")
async def test_rules(
    text: str = "Google and Microsoft use AI and machine learning for email services."
):
    """Test translation rules on sample text"""
    # Translate
    translations = translate_text_batch([text], "English", "Hindi")
    
    return {
        "original": text,
        "translated": translations[0],
        "rules_applied": {
            "transliterated": [word for word in text.split() if word.lower() in TRANSLITERATE_WORDS_LOWER],
            "kept_original": [word for word in text.split() if word.lower() in KEEP_ORIGINAL_WORDS_LOWER]
        }
    }
